from openpyxl import load_workbook
import random


class post(object):
    def __init__(self):
        wb = load_workbook('ASIN.xlsx')
        ws1 = wb['main']
        ws2 = wb['others']
        asin = []
        asin_o = []
        for row in ws1.rows:
            for cell in row:
                asin.append(cell.value)
        for row in ws2.rows:
            for cell in row:
                asin_o.append(cell.value)
        self.asin = asin
        self.asin_o = asin_o
        #初始化count 为0
        if ws1.max_column == 1:



    def random_choose(self, asin_list):
        get_asin = random.sample(asin_list,5)
        return get_asin

    def get_asin(self):
        get_asin = self.random_choose(self.asin)
        get_asin_o = self.random_choose(self.asin_o)
        return ",".join(get_asin+get_asin_o)


if __name__ == '__main__':
    test = post()
    print(test.get_asin())



